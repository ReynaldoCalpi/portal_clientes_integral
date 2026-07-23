import streamlit as st
import pandas as pd
from datetime import datetime
import json
import os
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de la página
st.set_page_config(
    page_title="Portal de Clientes - RI Consultores",
    page_icon="📊",
    layout="wide"
)

# --- Configuración de Persistencia en Disco ---
DB_FILE = "submissions_db.json"
PAYROLL_DB_FILE = "payroll_db.json"
EMPLOYEE_DB_FILE = "employees_db.json"
UPLOAD_DIR = "uploaded_files"

if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)

def load_json_db(file_path):
    if os.path.exists(file_path):
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return []
    return []

def save_json_db(file_path, data):
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def save_files_to_folder(file_list, client_name, periodo_str, category):
    saved_files_info = []
    if not file_list:
        return saved_files_info
        
    safe_client = client_name.replace(" ", "_").replace(".", "")
    safe_periodo = periodo_str.replace(" ", "_").replace("—", "-")
    folder_path = os.path.join(UPLOAD_DIR, safe_client, safe_periodo, category)
    os.makedirs(folder_path, exist_ok=True)
    
    for file_obj in file_list:
        file_path = os.path.join(folder_path, file_obj.name)
        file_obj.seek(0)
        with open(file_path, "wb") as f:
            f.write(file_obj.getbuffer())
        saved_files_info.append({
            "name": file_obj.name,
            "path": file_path
        })
    return saved_files_info

def create_zip_buffer(json_list, pdf_list):
    """Crea un archivo ZIP en memoria con los JSONs y PDFs proporcionados."""
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file_info in (json_list or []):
            if os.path.exists(file_info['path']):
                zip_file.write(file_info['path'], arcname=file_info['name'])
        for file_info in (pdf_list or []):
            if os.path.exists(file_info['path']):
                zip_file.write(file_info['path'], arcname=file_info['name'])
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def create_excel_payroll_buffer(client_name, periodo_str, items_list):
    """Genera un archivo Excel (.xlsx) con diseño profesional, formatos de moneda y fórmulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla Quincenal"
    ws.views.sheetView[0].showGridLines = True

    HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Arial", size=16, bold=True, color="1B365D")
    SUBTITLE_FONT = Font(name="Arial", size=11, italic=True, color="555555")
    DATA_FONT = Font(name="Arial", size=10)
    TOTAL_FONT = Font(name="Arial", size=10, bold=True)
    TOTAL_FILL = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000'),
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3')
    )

    ws['A1'] = "RI CONSULTORES — PLANILLA DE PAGO QUINCENAL"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Periodo: {periodo_str} | Empresa: {client_name}"
    ws['A2'].font = SUBTITLE_FONT

    ws.append([]) # Fila 3 en blanco

    headers = [
        "No.", "Empleado", "DUI", "Sueldo Quincenal", "Comisiones", 
        "Hrs Diurnas (200%)", "Hrs Nocturnas (225%)", "Total Gravable", 
        "ISSS (3%)", "AFP (7.25%)", "Renta", "Otras Deduc.", 
        "Código Fiscal", "Líquido a Pagar"
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    currency_format = "$#,##0.00"

    def clean_val(v):
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v.replace("$", "").replace(",", ""))
            except:
                return 0.0
        return 0.0

    for idx, item in enumerate(items_list, start=1):
        sq = clean_val(item.get("SUELDO QUINCENAL", 0))
        com = clean_val(item.get("COMISIONES", 0))
        h_d = clean_val(item.get("HRS DIURNAS (200%)", 0))
        h_n = clean_val(item.get("HRS NOCTURNAS (225%)", 0))
        tot_g = clean_val(item.get("TOTAL GRAVABLE", 0))
        isss = clean_val(item.get("ISSS (3%)", 0))
        afp = clean_val(item.get("AFP (7.25%)", 0))
        renta = clean_val(item.get("RENTA", 0))
        otras = clean_val(item.get("OTRAS DEDUCCIONES", 0))
        liq = clean_val(item.get("LÍQUIDO A PAGAR", 0))

        row_data = [
            idx,
            item.get("EMPLEADO", ""),
            item.get("DUI", ""),
            sq, com, h_d, h_n, tot_g, isss, afp, renta, otras,
            item.get("CÓDIGO FISCAL", "CÓDIGO 01"),
            liq
        ]
        ws.append(row_data)
        row_idx = 4 + idx
        is_zebra = (row_idx % 2 == 0)
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.border = thin_border
            if is_zebra:
                cell.fill = ZEBRA_FILL
            
            if col_idx in [1]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [2, 3, 13]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = currency_format
        ws.row_dimensions[row_idx].height = 20

    start_r = 5
    end_r = 4 + len(items_list)
    total_row_idx = end_r + 1

    totals_formula = [
        "TOTALES", "", "",
        f"=SUM(D{start_r}:D{end_r})",
        f"=SUM(E{start_r}:E{end_r})",
        f"=SUM(F{start_r}:F{end_r})",
        f"=SUM(G{start_r}:G{end_r})",
        f"=SUM(H{start_r}:H{end_r})",
        f"=SUM(I{start_r}:I{end_r})",
        f"=SUM(J{start_r}:J{end_r})",
        f"=SUM(K{start_r}:K{end_r})",
        f"=SUM(L{start_r}:L{end_r})",
        "",
        f"=SUM(N{start_r}:N{end_r})"
    ]

    ws.append(totals_formula)
    ws.row_dimensions[total_row_idx].height = 22

    for col_idx in range(1, len(totals_formula) + 1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = total_border
        if col_idx in [4, 5, 6, 7, 8, 9, 10, 11, 12, 14]:
            cell.number_format = currency_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 4:
                continue
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = "$999,999.00"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer.getvalue()

def calcular_empleado_quincenal(salario_mensual, comisiones, h_diurnas, h_nocturnas, otras_deducciones, tipo_regimen):
    """Calcula subtotal, ISSS, AFP, Renta quincenal según fórmula exacta de tramos y líquido a pagar."""
    tarifa_hora = (salario_mensual / 30.0) / 8.0
    
    pago_diurnas = h_diurnas * tarifa_hora * 2.0
    pago_nocturnas = h_nocturnas * tarifa_hora * 2.25
    
    salario_quincenal_base = salario_mensual / 2.0
    total_gravable = salario_quincenal_base + comisiones + pago_diurnas + pago_nocturnas
    
    # ISSS Empleado: 3% con tope quincenal de $15.00
    isss = min(total_gravable * 0.03, 15.00)
    
    # AFP Empleado: 7.25%
    afp = min(total_gravable * 0.0725, 7045.06 / 2.0)
    
    # Base gravable para Renta
    base_renta = total_gravable - isss - afp
    if base_renta < 0:
        base_renta = 0.0
        
    # Renta Quincenal exacta según fórmula de tramos de Ley
    if tipo_regimen == "Eventual (10%)":
        renta = total_gravable * 0.10
    elif tipo_regimen == "Exento / Código 60":
        renta = 0.0
    else:
        if base_renta <= 275.00:
            renta = 0.0
        elif base_renta <= 447.62:
            renta = ((base_renta - 275.00) * 0.10) + 8.83
        elif base_renta <= 1019.05:
            renta = ((base_renta - 447.62) * 0.20) + 30.00
        else:
            renta = ((base_renta - 1019.05) * 0.30) + 144.28
            
    if tipo_regimen == "Exento / Código 60" or (tipo_regimen == "Cálculo por Tramos de Ley" and base_renta <= 275.00):
        codigo_fiscal = "CÓDIGO 60"
    elif tipo_regimen == "Eventual (10%)":
        codigo_fiscal = "EVENTUAL 10%"
    else:
        codigo_fiscal = "CÓDIGO 01"
        
    liquido = total_gravable - isss - afp - renta - otras_deducciones
    
    return {
        "salario_quincenal": salario_quincenal_base,
        "comisiones": comisiones,
        "pago_diurnas": pago_diurnas,
        "pago_nocturnas": pago_nocturnas,
        "total_gravable": total_gravable,
        "isss": isss,
        "afp": afp,
        "renta": renta,
        "otras_deducciones": otras_deducciones,
        "codigo_fiscal": codigo_fiscal,
        "liquido": liquido
    }

# --- Inicialización de Estados de Sesión ---
if "logged_in" not in st.session_state: st.session_state.logged_in = False
if "user_role" not in st.session_state: st.session_state.user_role = None
if "username" not in st.session_state: st.session_state.username = ""
if "user_id" not in st.session_state: st.session_state.user_id = ""
if "clients_db" not in st.session_state: st.session_state.clients_db = {}

official_clients = {
    "admin": {"password": "admin123", "role": "admin", "name": "Administrador General"},
    "soluciones_503": {"password": "sol503_2026", "role": "client", "name": "Soluciones 503 S.A.S. de C.V"},
    "distribuidora_libertad": {"password": "libertad_2026", "role": "client", "name": "Distribuidora Libertad"},
    "leftech": {"password": "leftech_2026", "role": "client", "name": "Leftech"},
    "cedillo": {"password": "cedillo_2026", "role": "client", "name": "Cedillo"},
    "mercadito_rosa": {"password": "rosa_2026", "role": "client", "name": "Mercadito Rosa de Saron AC"}
}

for k, v in official_clients.items():
    st.session_state.clients_db[k] = v

# --- Pantalla de Login ---
def login_screen():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.title("🔐 RI Consultores")
        st.markdown("### Portal de Gestión Documental y Fiscal")
        with st.form("login_form"):
            username = st.text_input("USUARIO")
            password = st.text_input("CONTRASEÑA", type="password")
            submit = st.form_submit_button("INICIAR SESIÓN", use_container_width=True)
            if submit:
                user_key = username.strip().lower()
                if user_key in st.session_state.clients_db and st.session_state.clients_db[user_key]["password"] == password:
                    st.session_state.logged_in = True
                    st.session_state.user_role = st.session_state.clients_db[user_key]["role"]
                    st.session_state.username = st.session_state.clients_db[user_key]["name"]
                    st.session_state.user_id = user_key
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos.")

# --- Panel de Administración ---
def admin_dashboard():
    st.title("🎛️ PANEL DE CONTROL - ADMINISTRADOR")
    st.markdown("Supervisa cumplimiento fiscal, documentos, notas aclaratorias y planillas de sueldos enviadas.")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📋 DOCUMENTOS Y NOTAS", "💼 AUDITORÍA DE PLANILLAS", "➕ CREAR USUARIO", "👥 CLIENTES"])
    
    with tab1:
        st.subheader("CONTROL DE RECEPCIÓN DE DOCUMENTOS Y NOTAS")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            filtro_mes = st.selectbox("MES", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], index=5, key="admin_mes")
        with col_f2:
            filtro_anio = st.selectbox("AÑO", [2026, 2025], index=0, key="admin_anio")
            
        periodo_seleccionado = f"{filtro_mes} {filtro_anio}"
        all_submissions = load_json_db(DB_FILE)
        envios_periodo = [s for s in all_submissions if filtro_mes in s["periodo"] and str(filtro_anio) in s["periodo"]]
        
        if envios_periodo:
            for idx, envio in enumerate(envios_periodo):
                with st.expander(f"📁 {envio['client']} — Periodo: {envio['periodo']} — Entregado el {envio['fecha']}"):
                    if envio.get('notes'):
                        st.info(f"**📝 Notas / Aclaraciones del Cliente:**\n\n{envio['notes']}")
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        if envio.get('sales_json_list') or envio.get('sales_pdf_list'):
                            zip_b = create_zip_buffer(envio.get('sales_json_list'), envio.get('sales_pdf_list'))
                            st.download_button("📦 DESCARGAR VENTAS (ZIP)", data=zip_b, file_name=f"Ventas_{envio['client']}.zip", key=f"z_s_{idx}")
                        else:
                            st.text("Sin ventas")
                    with col_d2:
                        if envio.get('purch_json_list') or envio.get('purch_pdf_list'):
                            zip_b = create_zip_buffer(envio.get('purch_json_list'), envio.get('purch_pdf_list'))
                            st.download_button("📦 DESCARGAR COMPRAS (ZIP)", data=zip_b, file_name=f"Compras_{envio['client']}.zip", key=f"z_p_{idx}")
                        else:
                            st.text("Sin compras")
        else:
            st.info(f"No hay documentos para {periodo_seleccionado}.")

    with tab2:
        st.subheader("💼 AUDITORÍA DE PLANILLAS QUINCENALES ENVIADAS")
        payrolls = load_json_db(PAYROLL_DB_FILE)
        
        if payrolls:
            for p_idx, payroll in enumerate(payrolls):
                estado_str = "🟢 ENVIADA" if payroll.get("enviado", False) else "🟠 BORRADOR"
                with st.expander(f"🏢 EMPRESA: {payroll['client_name']} — PERIODO: {payroll['periodo']} [{estado_str}]"):
                    df_p = pd.DataFrame(payroll['items'])
                    st.dataframe(df_p, use_container_width=True)
                    
                    # Botón de descarga Excel en Admin
                    excel_data = create_excel_payroll_buffer(payroll['client_name'], payroll['periodo'], payroll['items'])
                    safe_c_name = payroll['client_name'].replace(" ", "_")
                    safe_p_name = payroll['periodo'].replace(" ", "_").replace("—", "-")
                    st.download_button(
                        label="📊 DESCARGAR PLANILLA EN EXCEL (PROFESIONAL)",
                        data=excel_data,
                        file_name=f"Planilla_{safe_c_name}_{safe_p_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"adm_excel_{p_idx}"
                    )
                    
                    if st.button(f"🗑️ [ADMIN] ELIMINAR ESTA PLANILLA", key=f"adm_del_{p_idx}", type="primary"):
                        updated_payrolls = [p for p in payrolls if not (p.get("user_id") == payroll['user_id'] and p.get("periodo") == payroll['periodo'])]
                        save_json_db(PAYROLL_DB_FILE, updated_payrolls)
                        st.success("Planilla eliminada exitosamente del sistema.")
                        st.rerun()
        else:
            st.info("Aún no hay planillas registradas en el sistema.")

    with tab3:
        st.subheader("REGISTRAR NUEVO CLIENTE")
        with st.form("new_client_form"):
            new_id = st.text_input("ID DE USUARIO (ej. empresa_xyz)").strip().lower()
            c_name = st.text_input("NOMBRE O RAZÓN SOCIAL")
            t_pass = st.text_input("CONTRASEÑA TEMPORAL", type="password")
            if st.form_submit_button("REGISTRAR"):
                if new_id and c_name and t_pass:
                    st.session_state.clients_db[new_id] = {"password": t_pass, "role": "client", "name": c_name}
                    st.success(f"Cliente {c_name} registrado.")
                else:
                    st.warning("Complete todos los campos.")

    with tab4:
        client_accounts = [{"USUARIO ID": k, "NOMBRE": v["name"]} for k, v in st.session_state.clients_db.items() if v["role"] == "client"]
        st.dataframe(pd.DataFrame(client_accounts), use_container_width=True)

# --- Panel del Cliente ---
def client_dashboard():
    st.title(f"📁 PORTAL DE CONTRIBUYENTE — {st.session_state.username}")
    st.markdown("Gestión documental, notas aclaratorias y generación de planillas fiscales quincenales.")
    
    col_p1, col_p2, col_p3 = st.columns(3)
    with col_p1:
        mes = st.selectbox("MES FISCAL", ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"], index=5, key="c_mes")
    with col_p2:
        anio = st.selectbox("AÑO FISCAL", [2026, 2025], index=0, key="c_anio")
    with col_p3:
        quincena = st.selectbox("QUINCENA", ["Primera Quincena (Del 1 al 15)", "Segunda Quincena (Del 16 al 30/31)"], key="c_quincena")
        
    periodo_str = f"{mes} {anio} — {quincena}"
    current_user_id = st.session_state.get("user_id", st.session_state.username)
    
    client_tab1, client_tab2, client_tab3 = st.tabs(["📤 CARGA DOCUMENTAL Y NOTAS", "💼 GENERADOR DE PLANILLAS", "📊 HISTORIAL Y RESUMEN"])
    
    with client_tab1:
        with st.form("upload_form"):
            col_v, col_c = st.columns(2)
            with col_v:
                st.subheader("📈 VENTAS")
                sales_json = st.file_uploader("JSON de Ventas", type=["json"], accept_multiple_files=True, key="s_j")
                sales_pdf = st.file_uploader("PDFs / ZIP de Ventas", type=["pdf", "zip"], accept_multiple_files=True, key="s_p")
            with col_c:
                st.subheader("📉 COMPRAS")
                purch_json = st.file_uploader("JSON de Compras", type=["json"], accept_multiple_files=True, key="p_j")
                purch_pdf = st.file_uploader("PDFs / ZIP de Compras", type=["pdf", "zip"], accept_multiple_files=True, key="p_p")
                
            st.divider()
            st.subheader("📝 NOTAS ACLARATORIAS Y OBSERVACIONES DEL MES")
            client_notes = st.text_area("Detalle aclaraciones sobre anulaciones, notas de crédito o situaciones especiales:", key="c_notes")
            
            if st.form_submit_button("🚀 ENVIAR DOCUMENTACIÓN Y NOTAS", use_container_width=True):
                if sales_json or purch_json:
                    s_js = save_files_to_folder(sales_json, st.session_state.username, periodo_str, "sales_json")
                    s_ps = save_files_to_folder(sales_pdf, st.session_state.username, periodo_str, "sales_pdf")
                    p_js = save_files_to_folder(purch_json, st.session_state.username, periodo_str, "purch_json")
                    p_ps = save_files_to_folder(purch_pdf, st.session_state.username, periodo_str, "purch_pdf")
                    
                    sub = {
                        "user_id": current_user_id, "client": st.session_state.username,
                        "periodo": periodo_str, "sales_json_list": s_js, "sales_pdf_list": s_ps,
                        "purch_json_list": p_js, "purch_pdf_list": p_ps, "notes": client_notes,
                        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M")
                    }
                    all_s = load_json_db(DB_FILE)
                    all_s.append(sub)
                    save_json_db(DB_FILE, all_s)
                    st.success("¡Documentación enviada con éxito!")
                    st.rerun()
                else:
                    st.warning("Adjunte al menos un JSON principal.")

    with client_tab2:
        st.subheader("💼 MANTENIMIENTO DE PERSONAL Y PLANILLA QUINCENAL")
        st.markdown("Calcula automáticamente ISSS (3%), AFP (7.25%), Horas Extras (200% y 225%) y Renta Quincenal por tramos exactos.")
        
        all_emps = load_json_db(EMPLOYEE_DB_FILE)
        my_emps = [e for e in all_emps if e.get("user_id") == current_user_id]
        
        with st.expander("➕ REGISTRAR / ACTUALIZAR EMPLEADO EN BASE HISTÓRICA"):
            with st.form("emp_form"):
                e_nombre = st.text_input("NOMBRE COMPLETO DEL EMPLEADO")
                e_dui = st.text_input("DUI / IDENTIFICACIÓN")
                e_salario = st.number_input("SALARIO BASE MENSUAL ($)", min_value=0.0, value=500.0, step=10.0)
                e_regimen = st.selectbox("RÉGIMEN DE RETENCIÓN DE RENTA", ["Cálculo por Tramos de Ley", "Exento / Código 60", "Eventual (10%)"])
                
                if st.form_submit_button("GUARDAR EMPLEADO"):
                    if e_nombre and e_dui:
                        new_list = [e for e in all_emps if not (e.get("user_id") == current_user_id and e.get("dui") == e_dui)]
                        new_list.append({
                            "user_id": current_user_id, "nombre": e_nombre, "dui": e_dui,
                            "salario_base": e_salario, "regimen": e_regimen
                        })
                        save_json_db(EMPLOYEE_DB_FILE, new_list)
                        st.success(f"Empleado {e_nombre} guardado correctamente.")
                        st.rerun()
                    else:
                        st.warning("Ingrese nombre y DUI.")
                        
        if my_emps:
            st.markdown(f"**Empleados Registrados en Base Histórica ({len(my_emps)}):**")
            df_empleados = pd.DataFrame(my_emps)[['nombre', 'dui', 'salario_base', 'regimen']]
            df_empleados.columns = ["NOMBRE", "DUI", "SALARIO BASE ($)", "RÉGIMEN"]
            df_empleados["SALARIO BASE ($)"] = df_empleados["SALARIO BASE ($)"].apply(lambda x: f"${x:,.2f}")
            st.dataframe(df_empleados, use_container_width=True)
            
            st.markdown("---")
            st.subheader(f"GESTIÓN DE PLANILLA PARA EL PERIODO: {periodo_str}")
            
            all_payrolls = load_json_db(PAYROLL_DB_FILE)
            existing_payroll = next((p for p in all_payrolls if p.get("user_id") == current_user_id and p.get("periodo") == periodo_str), None)
            
            if existing_payroll:
                is_enviado = existing_payroll.get("enviado", False)
                
                st.dataframe(pd.DataFrame(existing_payroll['items']), use_container_width=True)
                
                # Botón de Descarga Excel Profesional para el Cliente
                excel_data = create_excel_payroll_buffer(st.session_state.username, periodo_str, existing_payroll['items'])
                safe_p_name = periodo_str.replace(" ", "_").replace("—", "-")
                st.download_button(
                    label="📊 DESCARGAR PLANILLA EN EXCEL (FORMATO PROFESIONAL)",
                    data=excel_data,
                    file_name=f"Planilla_{st.session_state.username.replace(' ', '_')}_{safe_p_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                if is_enviado:
                    st.info("🔒 **Planilla enviada oficialmente a RI Consultores.** Ya no puede ser modificada ni eliminada por ti.")
                else:
                    st.warning("⚠️ Tienes esta planilla guardada como **BORRADOR** (Aún no enviada).")
                    col_b1, col_b2 = st.columns(2)
                    with col_b1:
                        if st.button("🗑️ ELIMINAR BORRADOR DE PLANILLA", type="secondary"):
                            all_payrolls = [p for p in all_payrolls if not (p.get("user_id") == current_user_id and p.get("periodo") == periodo_str)]
                            save_json_db(PAYROLL_DB_FILE, all_payrolls)
                            st.success("Borrador eliminado correctamente.")
                            st.rerun()
                    with col_b2:
                        if st.button("🚀 ENVIAR PLANILLA A RI CONSULTORES", type="primary"):
                            for p in all_payrolls:
                                if p.get("user_id") == current_user_id and p.get("periodo") == periodo_str:
                                    p["enviado"] = True
                                    p["fecha_envio"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                            save_json_db(PAYROLL_DB_FILE, all_payrolls)
                            st.success("¡Planilla enviada a RI Consultores exitosamente!")
                            st.rerun()
            else:
                with st.form("payroll_generation_form"):
                    st.markdown("Ingrese **Comisiones**, **Horas Extras Diurnas (200%)**, **Horas Nocturnas (225%)** y **Otras Deducciones**:")
                    
                    emp_inputs = {}
                    for emp in my_emps:
                        sal_quincenal_base = emp['salario_base'] / 2.0
                        st.markdown(f"**{emp['nombre']}** (Base Quincenal: **${sal_quincenal_base:,.2f}**)")
                        
                        col_c1, col_c2, col_c3, col_c4 = st.columns(4)
                        with col_c1:
                            comisiones = st.number_input(f"Comisiones ({emp['dui']})", min_value=0.0, value=0.0, step=10.0, key=f"com_{emp['dui']}")
                        with col_c2:
                            h_diurnas = st.number_input(f"Hrs Diurnas 200% ({emp['dui']})", min_value=0.0, value=0.0, step=0.5, key=f"hd_{emp['dui']}")
                        with col_c3:
                            h_nocturnas = st.number_input(f"Hrs Nocturnas 225% ({emp['dui']})", min_value=0.0, value=0.0, step=0.5, key=f"hn_{emp['dui']}")
                        with col_c4:
                            otras_ded = st.number_input(f"Otras Deduc. ({emp['dui']})", min_value=0.0, value=0.0, step=5.0, key=f"od_{emp['dui']}")
                            
                        emp_inputs[emp['dui']] = {
                            "comisiones": comisiones,
                            "diurnas": h_diurnas,
                            "nocturnas": h_nocturnas,
                            "otras_deducciones": otras_ded
                        }
                        st.divider()
                    
                    submitted_draft = st.form_submit_button("💾 CALCULAR Y GUARDAR COMO BORRADOR", type="primary")
                    if submitted_draft:
                        planilla_items = []
                        for emp in my_emps:
                            inputs = emp_inputs[emp['dui']]
                            res = calcular_empleado_quincenal(
                                emp['salario_base'],
                                inputs["comisiones"],
                                inputs["diurnas"],
                                inputs["nocturnas"],
                                inputs["otras_deducciones"],
                                emp['regimen']
                            )
                            
                            planilla_items.append({
                                "EMPLEADO": emp['nombre'],
                                "DUI": emp['dui'],
                                "SUELDO QUINCENAL": f"${res['salario_quincenal']:,.2f}",
                                "COMISIONES": f"${res['comisiones']:,.2f}",
                                "HRS DIURNAS (200%)": f"${res['pago_diurnas']:,.2f}",
                                "HRS NOCTURNAS (225%)": f"${res['pago_nocturnas']:,.2f}",
                                "TOTAL GRAVABLE": f"${res['total_gravable']:,.2f}",
                                "ISSS (3%)": f"${res['isss']:,.2f}",
                                "AFP (7.25%)": f"${res['afp']:,.2f}",
                                "RENTA": f"${res['renta']:,.2f}",
                                "OTRAS DEDUCCIONES": f"${res['otras_deducciones']:,.2f}",
                                "CÓDIGO FISCAL": res['codigo_fiscal'],
                                "LÍQUIDO A PAGAR": f"${res['liquido']:,.2f}"
                            })
                            
                        new_payroll_record = {
                            "user_id": current_user_id,
                            "client_name": st.session_state.username,
                            "periodo": periodo_str,
                            "enviado": False,
                            "fecha_creacion": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            "items": planilla_items
                        }
                        all_payrolls.append(new_payroll_record)
                        save_json_db(PAYROLL_DB_FILE, all_payrolls)
                        st.success("¡Planilla guardada como borrador!")
                        st.rerun()
        else:
            st.info("Registra al menos un empleado en la base histórica superior.")

    with client_tab3:
        st.subheader("📊 HISTORIAL DE PLANILLAS QUINCENALES")
        all_p = load_json_db(PAYROLL_DB_FILE)
        my_p = [p for p in all_p if p.get("user_id") == current_user_id]
        if my_p:
            for h_idx, p in enumerate(my_p):
                estado_txt = "ENVIADA A RI CONSULTORES" if p.get("enviado", False) else "BORRADOR"
                with st.expander(f"Periodo: {p['periodo']} — Estado: [{estado_txt}] (Creada: {p['fecha_creacion']})"):
                    st.dataframe(pd.DataFrame(p['items']), use_container_width=True)
                    
                    excel_data = create_excel_payroll_buffer(st.session_state.username, p['periodo'], p['items'])
                    safe_p_name = p['periodo'].replace(" ", "_").replace("—", "-")
                    st.download_button(
                        label="📊 DESCARGAR ESTA PLANILLA EN EXCEL",
                        data=excel_data,
                        file_name=f"Planilla_{st.session_state.username.replace(' ', '_')}_{safe_p_name}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"hist_excel_{h_idx}"
                    )
        else:
            st.warning("No hay planillas registradas todavía.")

# --- Control de Sesión ---
if not st.session_state.logged_in:
    login_screen()
else:
    with st.sidebar:
        st.write(f"Conectado como:\n**{st.session_state.username}**")
        st.divider()
        if st.button("Cerrar Sesión", type="primary"):
            st.session_state.logged_in = False
            st.session_state.user_role = None
            st.session_state.username = ""
            st.session_state.user_id = ""
            st.rerun()
            
    if st.session_state.user_role == "admin":
        admin_dashboard()
    elif st.session_state.user_role == "client":
        client_dashboard()
